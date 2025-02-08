import pandas as pd
pd.options.mode.chained_assignment = None
from openpyxl import load_workbook
import numpy as np
from scipy.optimize import minimize
import time
import itertools
import pprint

#============================================================================
def load_table_into_dataframe(file_path, sheet_name, table_name):
    print("Loading", file_path)
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    data = ws.tables[table_name]
    
    # Load data directly into DataFrame
    data_rows = [[cell.value for cell in row] for row in ws[data.ref]]
    df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
    
    print("Loaded", df.shape[0], "radiators")
    return df

#============================================================================
def create_test_rooms():
    test_rooms = [
        {'name': 'lounge', 'room_temperature': 21.5, 'min_wattage': 4000},
        {'name': 'bed 1', 'room_temperature': 21.5, 'min_wattage': 500},
        {'name': 'bed 2', 'room_temperature': 21.5, 'min_wattage': 1000},
        {'name': 'bed 3', 'room_temperature': 21.5, 'min_wattage': 800},
        {'name': 'bed 4', 'room_temperature': 21.5, 'min_wattage': 1000},
    ]
    return pd.DataFrame(test_rooms)

#============================================================================
def create_test_constraints():
    test_constraints = [
        {'name': 'lounge', 'location': 'Loc 1', 'Type': 'Modern', 'Height': 600, 'Length': 2000, 'Depth': 'K2', 'Labour Cost':  95.0, 'Existing Radiator': 'ModernxK2x1000x600'},
        {'name': 'lounge', 'location': 'Loc 2', 'Type': 'Modern', 'Height': 600, 'Length': 2000, 'Depth': 'K3', 'Labour Cost': 350.0, 'Existing Radiator': None},
        {'name': 'lounge', 'location': 'Loc 3', 'Type': 'Modern', 'Height': 600, 'Length':  600, 'Depth': 'K2', 'Labour Cost': 350.0, 'Existing Radiator': None},
        {'name': 'lounge', 'location': 'Loc 4', 'Type': 'Modern', 'Height': 600, 'Length':  600, 'Depth': 'K3', 'Labour Cost': 350.0, 'Existing Radiator': None},
        {'name': 'bed 1',  'location': 'Loc 1', 'Type': 'Modern', 'Height': 600, 'Length': 2000, 'Depth': 'K2', 'Labour Cost':  95.0, 'Existing Radiator': 'ModernxK1x1000x600'},
        {'name': 'bed 1',  'location': 'Loc 2', 'Type': 'Modern', 'Height': 600, 'Length': 2000, 'Depth': 'K3', 'Labour Cost': 350.0, 'Existing Radiator': None},
        {'name': 'bed 2',  'location': 'Loc 1', 'Type': 'Modern', 'Height': 900, 'Length': 2000, 'Depth': 'K3', 'Labour Cost':  95.0, 'Existing Radiator': 'ModernxK2x1400x600'},
        {'name': 'bed 3',  'location': 'Loc 1', 'Type': 'Modern', 'Height': 900, 'Length': 2000, 'Depth': 'K3', 'Labour Cost':  95.0, 'Existing Radiator': 'ModernxK1x1400x600'},
        {'name': 'bed 4',  'location': 'Loc 1', 'Type': 'Modern', 'Height': 900, 'Length': 2000, 'Depth': 'K3', 'Labour Cost':  95.0, 'Existing Radiator': 'ModernxK1x1800x600'}
    ]
    return pd.DataFrame(test_constraints)

#=======================================================================================================
class Radiator:
    radiator_depth_mm_data = {
        '10':  70,  # 1 panel 0 fins
        'P1':  70,
        'K1':  80, # 1 panel 1 fin
        '11':  80,
        'P+': 100, # 2 panels 1 fin
        '21': 100,
        'K2': 125, # 2 panels 2 fins
        '22': 125,
        'K3': 160, # 3 panels 3 fins
        '33': 160,
        '2 Col':  90,
        '3 Col': 115,
        '4 Col': 140,
        '5 Col': 175,
        '6 Col': 210,
        '7 Col': 245,
        None:   None,
    }

    @classmethod
    def radiator_depth_mm(cls, depth):
        return cls.radiator_depth_mm_data[depth]
        
    @classmethod
    def radiator_fits(cls, location, radiator):
        return (radiator['Type']   == location['Type'] and
                radiator['Height'] <= location['Height'] and
                radiator['Length'] <= location['Length'] and
                cls.radiator_depth_mm(radiator['Depth']) <= cls.radiator_depth_mm(location['Depth']))
    
    @classmethod
    def radiator_choices_at_location(cls, radiator_database, constraint):
        constraint_radiator_depth_mm = cls.radiator_depth_mm(constraint['Depth'])

        return radiator_database[(radiator_database['Type']   <= constraint['Type']) &
                                 (radiator_database['Length']   <= constraint['Length']) & 
                                 (radiator_database['Height']   <= constraint['Height']) &
                                 (radiator_database['Depth_mm'] <= constraint_radiator_depth_mm)]

#=======================================================================================================
class Home:
    def __init__(self, rooms, radiator_constraints, radiator_database):
        self.rooms = rooms
        self.radiator_constraints = radiator_constraints
        self.radiator_database = radiator_database
        self.add_radiator_depth_mm()

    def minimal_cost_radiators(self, flow_temperature):
        room_results = [self.minimal_cost_radiators_in_room(room, flow_temperature) for _, room in self.rooms.iterrows()]
        
        total_cost = self.total_costs(room_results)
        print("Total cost before moves:", total_cost)

        self.move_replaced_radiators(flow_temperature, room_results)

        total_cost = self.total_costs(room_results)
        print("Total cost after moves:", total_cost)

        return self.convert_results_to_dataframe(room_results)

    def add_radiator_depth_mm(self):
        self.radiator_database['Depth_mm'] = self.radiator_database.apply(lambda row: Radiator.radiator_depth_mm(row.Depth), axis = 1)

    def convert_results_to_dataframe(self, room_results):
        formatted_results = []
        for room in room_results:
            for location_name, location in room['locations'].items():
                location_constraint = self.find_location_constraint(room['room_name'], location_name).iloc[0]
                existing_rad_name = location_constraint['Existing Radiator']
                rad_status = self.radiator_change_status(existing_rad_name, location['Key'], location)

                formatted_result = {
                    'Room Name': room['room_name'],
                    'Location name': location_name,
                    'Originally': existing_rad_name,
                    'Proposed Radiator': location['Key'],
                    '£': location['£'],
                    'Labour Cost': location['Labour Cost'],
                    'Status': rad_status,
                    'Watts': location['w']
                }
                formatted_results.append(formatted_result)
        
        return pd.DataFrame(formatted_results)

    def radiator_change_status(self, original_name, new_name, location):
        if 'Status' in location and location['Status'] == 'Moved':
            return f"Moved:{location['From']['room_name']}:{location['From']['location']}"
        if new_name is None and original_name is not None:
            return 'Removed'
        if original_name == new_name:
            return 'Original' if original_name is not None else ''
        return 'Replaced'

    def find_location_constraint(self, room_name, location_name):
        return self.radiator_constraints[(self.radiator_constraints['name'] == room_name) & 
                                         (self.radiator_constraints['location'] == location_name)]

    def minimal_cost_radiators_in_room(self, room_df, flow_temperature):
        room_name = room_df['name']
        location_constraints = self.radiator_constraints[self.radiator_constraints['name'] == room_name]
        room = Room(room_df, location_constraints, self.radiator_database)
        room_result = room.minimal_cost_radiators(flow_temperature)
        room_result['room_name'] = room_name
        room_result['room_temperature'] = room_df['room_temperature']
        return room_result

    def total_costs(self, room_results):
        return sum(room_result['cost'] for room_result in room_results)

    def extract_replaced_radiators(self, room_results):
        replaced_rads = [rad for room_result in room_results for rad in room_result['replaced_radiators']]
        for rad in replaced_rads:
            rad['specification'] = self.find_radiator(self.radiator_database, rad['Key'])
        replaced_rads.sort(key=lambda rad: rad['specification']['£'])
        return replaced_rads

    def move_replaced_radiators(self, flow_temperature, room_results):
        replaced_rads = self.extract_replaced_radiators(room_results)
        new_radiators = self.find_new_radiators(room_results)

        for replacement_rad in replaced_rads:
            for new_radiator in new_radiators:
                location_constraint = self.find_constraint(new_radiator['room_name'], new_radiator['location_name'])
                result_location = self.find_location_in_room_results(room_results, location_constraint['name'], location_constraint['location'])

                if replacement_rad['specification']['w'] >= new_radiator['radiator']['w'] and \
                   Radiator.radiator_fits(location_constraint, replacement_rad['specification']) and \
                   result_location.get('Status') != 'Moved':
                    result_location['Status'] = 'Moved'
                    result_location['From'] = replacement_rad
                    result_location.update(self.find_radiator(self.radiator_database, replacement_rad['Key']))
                    result_location['£'] = 0.0
                    break

    def find_constraint(self, room_name, location_name):
        return self.radiator_constraints[(self.radiator_constraints['name'] == room_name) & 
                                         (self.radiator_constraints['location'] == location_name)].iloc[0]

    def find_location_in_room_results(self, room_results, room_name, location_name):
        return self.find_room_in_room_results(room_results, room_name)['locations'][location_name]

    def find_room_in_room_results(self, room_results, room_name):
        return next(room_result for room_result in room_results if room_result['room_name'] == room_name)

    def find_radiator(self, rad_db, key):
        return rad_db[rad_db['Key'] == key].iloc[0] if key is not None else None

    def find_new_radiators(self, room_results):
        new_radiators = []
        for room in room_results:
            for location_name, location in room['locations'].items():
                if location.get('Status') == 'New':
                    new_radiators.append({
                        'room_name': room['room_name'],
                        'location_name': location_name,
                        'radiator': location,
                        'room_temperature': room['room_temperature']
                    })
        new_radiators.sort(reverse=True, key=lambda rad: rad['radiator']['£'])
        return new_radiators

#=======================================================================================================
class Room:
    def __init__(self, room, location_constraints, radiator_database):
        self.room = room
        self.location_constraints = location_constraints
        self.radiator_database = radiator_database

    def minimal_cost_radiators(self, flow_temperature):
        t0 = time.time()
        self.pre_calculate_radiator_wattage_at_flow(flow_temperature)
        combos = self.all_combinations(flow_temperature)
        cost, rads = self.minimum_radiator_cost_combination(combos, self.location_constraints, self.room['min_wattage'])
        
        print("Time taken", round(time.time() - t0, 2), "s for", len(combos), "combinations")
        replaced_rads = self.replaced_radiators(self.room['name'], rads, self.location_constraints)
        return {'cost': cost, 'locations': rads, 'replaced_radiators': replaced_rads}

    def pre_calculate_radiator_wattage_at_flow(self, flow_temperature):
        factor = (flow_temperature - 2.5 - self.room['room_temperature']) / 50.0
        self.radiator_database['w'] = self.radiator_database['W @ dt 50'] * factor ** self.radiator_database['N']

    def all_combinations(self, flow_temperature):
        possible_rads_at_location = []
        for _, constraint in self.location_constraints.iterrows():
            rads = Radiator.radiator_choices_at_location(self.radiator_database, constraint)

            # Fix: Use .loc to assign values
            rads.loc[:, 'Labour Cost'] = constraint['Labour Cost']
            rads.loc[:, 'Status'] = 'New'
            
            rads = self.add_existing_radiators_to_possible_radiators(rads, constraint, flow_temperature)
            rads = self.add_no_radiator_to_possible_radiators(rads, constraint)
            
            rads.loc[:, 'Total £'] = rads['£'] + rads['Labour Cost']
            possible_rads_at_location.append(rads.to_dict('records'))
        
        return list(itertools.product(*possible_rads_at_location))
    
    def add_existing_radiators_to_possible_radiators(self, possible_rads, constraint, flow_temperature):
        if 'Existing Radiator' in constraint and isinstance(constraint['Existing Radiator'], str):
            existing_rad = self.find_radiator(constraint['Existing Radiator'])
            factor = (flow_temperature - 2.5 - self.room['room_temperature']) / 50.0
            watts_at_flow = existing_rad['W @ dt 50'] * factor ** existing_rad['N']
            return self.add_radiator(possible_rads, existing_rad['Key'], watts_at_flow, 0.0, 0.0, 'Original')
        return possible_rads

    def add_no_radiator_to_possible_radiators(self, possible_rads, constraint):
        return self.add_radiator(possible_rads, None, 0.0, 0.0, 0.0, 'No radiator')

    def add_radiator(self, rads, key, w, cost, labour_cost, existing):
        new_rad = {'Key': key, 'w': w, '£': cost, 'Labour Cost': labour_cost, 'Status': existing}
        return pd.concat([pd.DataFrame(new_rad, index=[0]), rads], ignore_index=True)

    def minimum_radiator_cost_combination(self, combos, constraints, min_wattage):
        min_cost = float('inf')
        min_rads = None
        location_names = constraints['location'].tolist()
        labour_costs = constraints['Labour Cost'].tolist()

        # speed critical, optimised loop
        for rads in combos:
            cost, watts = self.cost_of_all_radiators(rads, labour_costs)
            if watts > min_wattage:
                if cost < min_cost:
                    min_cost = cost
                    min_rads = dict(zip(location_names, rads))

        return min_cost, min_rads
    
    # speed critical, optimised loop   
    def cost_of_all_radiators(self, rads, labour_costs):
        costs = 0.0
        watts = 0.0

        for rad in rads:
            costs += rad['Total £']
            watts += rad['w']

        return [costs, watts]

    def replaced_radiators(self, room_name, new_radiators, constraints):
        return [
            {'room_name': room_name, 'location': loc['location'], 'Key': loc['Existing Radiator']}
            for _, loc in constraints.iterrows()
            if 'Existing Radiator' in loc and self.radiator_changed(loc['Existing Radiator'], new_radiators[loc['location']]['Key'])
        ]

    def radiator_changed(self, original, potential_new):
        return isinstance(original, str) and isinstance(potential_new, str) and original != potential_new
    
    def find_radiator(self, key):
        return self.radiator_database.loc[self.radiator_database['Key'] == key].iloc[0]

#============================================================================
file_path = 'Radiator Database.xlsx'
sheet_name = 'RadiatorDatabase'
table_name = 'RadiatorDatabase'
flow_temperatures = [55.0, 52.5, 50.0, 47.5, 45.0, 42.5, 40.0, 37.5, 35.0]

rad_db = load_table_into_dataframe(file_path, sheet_name, table_name)
test_rooms = create_test_rooms()
test_constraints = create_test_constraints()

t1 = time.time()
home = Home(test_rooms, test_constraints, rad_db)
for flow_temperature in [50.0]:
    print("Flow temperature", flow_temperature)
    results = home.minimal_cost_radiators(flow_temperature)
    print("=" * 100)
    print("Time taken", time.time() - t1)
    pprint.pprint(results)
    print("Totals:")
    pprint.pprint(results[['£', 'Labour Cost']].sum())